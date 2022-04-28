from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import copy
from decimal import Decimal

class TxRfFuncConfig:
    def __init__(self, config_file_add):
        self.config_file_add = config_file_add
        self.workbook = load_workbook(config_file_add)
        #self.workbook.active
        self.sheets = self.workbook.sheetnames
        self.load_test_config()
        with open(self.output_file_name, 'w') as file_out:
            with open(self.testDef_head, 'r') as finle_in:
                testDef_head = finle_in.readlines()
            file_out.writelines(testDef_head)
            file_out.write('\n')
        self.load_common_config()
        self.load_carrier_limit_table()
        self.carrier_sheets = self.workbook.sheetnames[4::]



    def load_test_config(self):
        worksheet = self.workbook['TestConfig']
        self.testDef_head = worksheet.cell(row=2, column=2).value
        self.output_file_name = worksheet.cell(row=3, column=2).value
        self.start_index = int(worksheet.cell(row=4, column=2).value)
        self.trdc_delay_short   = worksheet.cell(row=5, column=2).value
        self.trdc_delay_long    = worksheet.cell(row=6, column=2).value
        self.branch = worksheet.cell(row=7, column=2).value
        self.AXC = worksheet.cell(row=8, column=2).value

    def load_common_config(self):
        worksheet = self.workbook['CommonConfig']
        max_row = worksheet.max_row
        self.para_tap = []
        self.para_value = []
        for i in range(2, max_row+1):
            self.para_tap.append(worksheet.cell(row=i, column=1).value)
            self.para_value.append(worksheet.cell(row=i, column=2).value)
        self.para_value_txblock = copy.deepcopy(self.para_value)
        for i in range(len(self.para_value_txblock)):
            self.para_value_txblock[i] = '-'
        index = self.para_tap.index('delayAfterTxOn')
        self.para_value_txblock[index] = 0
        index = self.para_tap.index('doMeas')
        self.para_value_txblock[index] = 0
        index = self.para_tap.index('measTime')
        self.para_value_txblock[index] = 0

    def load_carrier_limit_table(self):
        worksheet = self.workbook['CarrierLimit']
        max_row = worksheet.max_row
        self.carrier_limit_table = []
        for i in range(2,max_row+1):
            self.carrier_limit_table.append(list(worksheet.iter_rows(i, i, worksheet.min_column, worksheet.max_column, True))[0])

    def load_carrier_config(self, worksheet_name):
        worksheet = self.workbook[worksheet_name]
        max_row = worksheet.max_row
        current_row = 1
        carriers_config = []
        while current_row < max_row:
            row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            current_row = current_row + 1
            '''
            All the carrier configs must start with * at the first col
            '''
            if row_data[0] == r'*':
                print(row_data[1])
                row_data = list(filter(None, row_data))
                carrier_dict={}
                carrier_dict['commont'] = row_data[1]
                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                row_data = list(filter(None, row_data))
                current_row = current_row + 1
                carrier_dict['carrier_type'] = row_data

                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                row_data = list(filter(None, row_data))
                current_row = current_row + 1
                carrier_dict['test_mod'] = row_data

                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                row_data = list(filter(None, row_data))
                current_row = current_row + 1
                carrier_dict['power_backoff'] = row_data

                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]

                carrier_dict['freq'] = []
                while(row_data[0]==None and row_data[1]):
                    row_data = list(filter(None, row_data))
                    carrier_dict['freq'].append(row_data)
                    current_row = current_row + 1
                    row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column,True))[0]
                carriers_config.append(carrier_dict)
        return carriers_config

    def gen_carrier_para(self, carrier_dict):
        para_value_carrier = copy.deepcopy(self.para_value)
        spurEm=0
        carrPwr=0
        delayAfterTxOn=self.trdc_delay_short
        pwrAcc=99
        evmMax=99
        evmMaxIOT=99
        RBW='1000k'
        limitTable='BC1_NR2'

        for i in range(len(carrier_dict['carrier_type'])):
            for config in self.carrier_limit_table:
                if carrier_dict['carrier_type'][i] == config[0] and carrier_dict['test_mod'][i] == config[1] and carrier_dict['power_backoff'][i] == config[2]:
                    if config[2] == '0dB' or config[2] == '-0dB':
                        spurEm = 1
                        delayAfterTxOn = self.trdc_delay_long
                    if len(carrier_dict['carrier_type']) ==1:
                        if config[3] != 'NA':
                            if config[0] == 'L4':
                                evmMaxIOT = min(evmMaxIOT, float(config[3]))
                            evmMax = min(evmMax, float(config[3]))
                        if config[5] != 'NA':
                            carrPwr = 1
                            pwrAcc = min(pwrAcc, float(config[5]))
                    else:
                        if config[4] != 'NA':
                            if config[0] == 'L4':
                                evmMaxIOT = min(evmMaxIOT, float(config[4]))
                            evmMax = min(evmMax, float(config[4]))
                        if config[6] != 'NA':
                            carrPwr = 1
                            pwrAcc = min(pwrAcc, float(config[6]))
                    RBW = '{}k'.format(min(int(RBW.strip('k')), int(config[7].strip('k'))))
                    if config[8] != 'BC1_NR2':
                        limitTable = config[8]

        index = self.para_tap.index('spurEm')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = spurEm

        index = self.para_tap.index('carrPwr')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = carrPwr

        index = self.para_tap.index('delayAfterTxOn')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = delayAfterTxOn

        index = self.para_tap.index('pwrAcc')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = pwrAcc

        index = self.para_tap.index('evmMax')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = evmMax

        index = self.para_tap.index('evmMaxIOT')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = evmMaxIOT

        index = self.para_tap.index('RBW')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = RBW

        index = self.para_tap.index('limitTable')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = limitTable
        return para_value_carrier

    def gen_carriers_TxRfFunc(self, worksheet_name):
        carriers_config = self.load_carrier_config(worksheet_name)
        for carrier_config in carriers_config:
            print('Generating TX_RF_FUNC_{}'.format(self.start_index))
            with open(self.output_file_name, 'a') as f:
                f.write('\n*{}\n'.format(carrier_config['commont']))
                f.write('#TX_RF_FUNC_{:<10}chConf'.format(self.start_index, ''))
                f.write('{:<360}'.format(''))
                for i in self.para_tap:
                    f.write('{:<20}'.format(i))
                f.write('\n*----------------------------------------------------------------------------------------------------------------------------------\n')
                para_value_carrier = self.gen_carrier_para(carrier_config)
                for freq in carrier_config['freq']:
                    f.write('{:<28}'.format(''))
                    carrier_setup_string = ''
                    for i in range(len(freq)):
                        carrier_setup_string += '<{}_{}_{}_Dd_MOD{}_P{}_PROTtrdcsetup7_AXC{}bit>;'.format(self.branch,carrier_config['carrier_type'][i], freq[i],carrier_config['test_mod'][i],carrier_config['power_backoff'][i], self.AXC)
                    f.write('{:<360} '.format(carrier_setup_string))


                    for i in para_value_carrier:
                        f.write('{:<20}'.format(i))
                    f.write('\n')
                    f.write('{:<389}'.format('<TRblock> '))
                    for i in self.para_value_txblock:
                        f.write('{:<20}'.format(i))
                    f.write('\n')

            print(self.start_index)
            self.start_index = self.start_index + 1

    def gen_carrier_combination(self):
        worksheet_in = self.workbook['CarrierSetup']

        max_row = worksheet_in.max_row
        freq_l = Decimal(worksheet_in.cell(row=1, column=2).value)
        freq_h = Decimal(worksheet_in.cell(row=2, column=2).value)
        freq_m = (freq_l + freq_h)/2
        full_bw = freq_h - freq_l
        offset = 7
        carrier_setup = []
        carrier_setup_narrow = []
        current_row = 1
        current_col = 1
        for row in range(offset, max_row+1):
            carrier_dict={}
            carrier_dict['carrier_name'] = worksheet_in.cell(row=row, column=1).value
            carrier_dict['carrier_bw'] = Decimal(worksheet_in.cell(row=row, column=2).value)
            carrier_dict['carrier_testmod'] = worksheet_in.cell(row=row, column=3).value
            carrier_dict['carrier_backoff'] = worksheet_in.cell(row=row, column=4).value
            if worksheet_in.cell(row=row, column=5).value =='1':
                carrier_setup_narrow.append(carrier_dict)
            else:
                carrier_setup.append(carrier_dict)

        #Single carrier
        worksheet_out = self.workbook.create_sheet("1 carrier")
        current_row = 1
        # Single carrier wide band
        for i1 in range(len(carrier_setup)):
            bw = carrier_setup[i1]['carrier_bw']
            worksheet_out.cell(row=current_row, column=1).value = '*'
            worksheet_out.cell(row=current_row, column=2).value = '{} {} {}'.format(carrier_setup[i1]['carrier_name'],carrier_setup[i1]['carrier_testmod'],carrier_setup[i1]['carrier_backoff'] )
            current_row += 1
            current_col = 2
            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i1]['carrier_name']
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i1]['carrier_testmod']
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i1]['carrier_backoff']
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l+bw/2)
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m)
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h-bw/2)
            current_row += 2
        # Single carrier narrow band
        for i1 in range(len(carrier_setup_narrow)):
            bw = carrier_setup_narrow[i1]['carrier_bw']
            worksheet_out.cell(row=current_row, column=1).value = '*'
            worksheet_out.cell(row=current_row, column=2).value = '{} {} {}'.format(carrier_setup_narrow[i1]['carrier_name'],carrier_setup_narrow[i1]['carrier_testmod'],carrier_setup_narrow[i1]['carrier_backoff'])
            current_row += 1
            current_col = 2
            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_backoff']
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l + bw / 2)
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m)
            current_row += 1
            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h - bw / 2)
            current_row += 2
        self.workbook.save(self.config_file_add)

        # 2 carriers
        # 2 wide band
        worksheet_out = self.workbook.create_sheet("2 Carrier")
        current_row = 1
        for i1 in range(len(carrier_setup)):
            bw1 = carrier_setup[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                if i2 >= i1 and carrier_setup[i1]['carrier_bw'] + carrier_setup[i2]['carrier_bw'] <= full_bw:

                    worksheet_out.cell(row=current_row, column=1).value = '*'
                    worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {}'.format(carrier_setup[i1]['carrier_name'],carrier_setup[i2]['carrier_name'],carrier_setup[i1]['carrier_testmod'],carrier_setup[i2]['carrier_testmod'],carrier_setup[i1]['carrier_backoff'],carrier_setup[i2]['carrier_backoff'] )
                    current_row += 1
                    current_col = 2
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i1]['carrier_name']
                    current_col+=1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2]['carrier_name']
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i1]['carrier_testmod']
                    current_col+=1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2]['carrier_testmod']
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i1]['carrier_backoff']
                    current_col+=1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2]['carrier_backoff']
                    current_col = 2
                    current_row += 1

                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l+bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h-bw2/2)
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h-bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l+bw2/2)
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m+bw2/2)
                    current_col = 2
                    current_row += 2

        # 2 narrow band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup_narrow)):
                bw2 = carrier_setup_narrow[i2]['carrier_bw']
                if i2 >= i1 and bw1 + bw2 <= full_bw:

                    worksheet_out.cell(row=current_row, column=1).value = '*'
                    worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {}'.format(carrier_setup_narrow[i1]['carrier_name'],carrier_setup_narrow[i2]['carrier_name'],carrier_setup_narrow[i1]['carrier_testmod'],carrier_setup_narrow[i2]['carrier_testmod'],carrier_setup_narrow[i1]['carrier_backoff'],carrier_setup_narrow[i2]['carrier_backoff'] )
                    current_row += 1
                    current_col = 2
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
                    current_col+=1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_name']
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
                    current_col+=1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_testmod']
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_backoff']
                    current_col+=1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_backoff']
                    current_col = 2
                    current_row += 1

                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-10+bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m+10-bw2/2)
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m+bw2/2)
                    current_col = 2
                    current_row += 2

        # 1 narrow band 1 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                if i2 >= i1 and bw1 + bw2 <= full_bw:
                    worksheet_out.cell(row=current_row, column=1).value = '*'
                    worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {}'.format(
                        carrier_setup_narrow[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],
                        carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],
                        carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'])
                    current_row += 1
                    current_col = 2
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                        'carrier_name']
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2][
                        'carrier_name']
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                        'carrier_testmod']
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2][
                        'carrier_testmod']
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                        'carrier_backoff']
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2][
                        'carrier_backoff']
                    current_col = 2
                    current_row += 1

                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l+bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h-bw2/2)
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h-bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l+bw2/2)
                    current_col = 2
                    current_row += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-bw1/2)
                    current_col += 1
                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m+bw2/2)
                    current_col = 2
                    current_row += 2
        self.workbook.save(self.config_file_add)

        # 3 carriers
        worksheet_out = self.workbook.create_sheet("3 Carrier")
        current_row = 1
        # 3 wide band
        for i1 in range(len(carrier_setup)):
            bw1 = carrier_setup[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    index=[i1, i2, i3]
                    if i2 >= i1 and i3 >=i2 and carrier_setup[i1]['carrier_bw'] + carrier_setup[i2]['carrier_bw'] + carrier_setup[i3]['carrier_bw']<= full_bw:
                        worksheet_out.cell(row=current_row, column=1).value = '*'
                        worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {}'.format(
                            carrier_setup[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],
                            carrier_setup[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],
                            carrier_setup[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'])
                        current_row += 1
                        current_col = 2
                        for i in index:
                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                            current_col += 1
                        current_col = 2
                        current_row += 1
                        for i in index:
                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                            current_col += 1
                        current_col = 2
                        current_row += 1
                        for i in index:
                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                            current_col += 1
                        current_col = 2
                        current_row += 1
                        #3 seperate carriers
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l + bw1 / 2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format((freq_l + bw1 + freq_h - bw3)/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h - bw3 / 2)
                        current_col = 2
                        current_row += 1
                        # 3 close carriers
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1+bw2/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1+bw2+bw3/2)
                        current_col = 2
                        current_row += 1

                        current_row += 2


        # 1 narrow band 2 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = Decimal(carrier_setup_narrow[i1]['carrier_bw'])
            for i2 in range(len(carrier_setup)):
                bw2 = Decimal(carrier_setup[i2]['carrier_bw'])
                for i3 in range(len(carrier_setup)):
                    bw3 = Decimal(carrier_setup[i3]['carrier_bw'])
                    index = [i1, i2, i3]
                    if i2 >= i1 and i3>=i2 and  bw1 + bw2 + bw3 <= full_bw:
                        worksheet_out.cell(row=current_row, column=1).value = '*'
                        worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {}'.format(
                            carrier_setup_narrow[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],
                            carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],
                            carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'],carrier_setup[i2]['carrier_backoff'])
                        current_row += 1
                        current_col = 2
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                            'carrier_name']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2][
                            'carrier_name']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i3][
                            'carrier_name']
                        current_col = 2
                        current_row += 1


                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                            'carrier_testmod']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2][
                            'carrier_testmod']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i3][
                            'carrier_testmod']
                        current_col = 2
                        current_row += 1

                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                            'carrier_backoff']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i2][
                            'carrier_backoff']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i3][
                            'carrier_backoff']
                        current_col = 2
                        current_row += 1


                        #3 seperate carriers
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l + bw1 / 2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format((freq_l + bw1 + freq_h - bw3)/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h - bw3 / 2)
                        current_col = 2
                        current_row += 1
                        # 3 close carriers
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1+bw2/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1+bw2+bw3/2)
                        current_col = 2
                        current_row += 1

                        current_row += 2

        # 2 narrow band 1 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup_narrow)):
                bw2 = carrier_setup_narrow[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    index = [i1, i2, i3]
                    if i2 >= i1 and i3>=i2 and  bw1 + bw2 + bw3 <= full_bw:
                        worksheet_out.cell(row=current_row, column=1).value = '*'
                        worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {}'.format(
                            carrier_setup_narrow[i1]['carrier_name'], carrier_setup_narrow[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],
                            carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup_narrow[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],
                            carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup_narrow[i2]['carrier_backoff'],carrier_setup[i2]['carrier_backoff'])
                        current_row += 1
                        current_col = 2
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                            'carrier_name']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2][
                            'carrier_name']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i3][
                            'carrier_name']
                        current_col = 2
                        current_row += 1


                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                            'carrier_testmod']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2][
                            'carrier_testmod']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i3][
                            'carrier_testmod']
                        current_col = 2
                        current_row += 1

                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                            'carrier_backoff']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2][
                            'carrier_backoff']
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i3][
                            'carrier_backoff']
                        current_col = 2
                        current_row += 1


                        #3 seperate carriers
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_l + bw1 / 2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format((freq_l + bw1 + bw2/2))
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_h - bw3 / 2)
                        current_col = 2
                        current_row += 1
                        # 3 close carriers
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1+bw2/2)
                        current_col += 1
                        worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3)/2+bw1+bw2+bw3/2)
                        current_col = 2
                        current_row += 1

                        current_row += 2

        self.workbook.save(self.config_file_add)
        # 4 carriers
        worksheet_out = self.workbook.create_sheet("4 Carrier")
        current_row = 1
        # 4 wide band
        for i1 in range(len(carrier_setup)):
            bw1 = carrier_setup[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        index=[i1, i2, i3,i4]
                        if i2 >= i1 and i3 >=i2 and i4>=i3 and bw1 + bw2 + bw3 + bw4 <= full_bw:
                            worksheet_out.cell(row=current_row, column=1).value = '*'
                            worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                carrier_setup[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],
                                carrier_setup[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],
                                carrier_setup[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'])
                            current_row += 1
                            current_col = 2
                            for i in index:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                current_col += 1
                            current_col = 2
                            current_row += 1
                            for i in index:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                current_col += 1
                            current_col = 2
                            current_row += 1
                            for i in index:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                current_col += 1
                            current_col = 2
                            current_row += 1
                            #4 seperate carriers
                            if bw1+bw2+bw3+bw4 < full_bw - 5:
                                freq_gap = (full_bw - bw1-bw2-bw3-bw4)/3
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw4 - freq_gap - bw3/2))
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw4/2)
                                current_col = 2
                                current_row += 1

                            # 4 close carriers
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1/2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1+bw2/2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1+bw2 + bw3/2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1+bw2+bw3 + bw4/2)
                            current_col = 2
                            current_row += 1

                            current_row += 2


        # 1 narrow band 3 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        index=[i1, i2, i3,i4]
                        if i2 >= i1 and i3 >=i2 and i4>=i3 and bw1 + bw2 + bw3 + bw4 <= full_bw:
                            worksheet_out.cell(row=current_row, column=1).value = '*'
                            worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                carrier_setup_narrow[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],
                                carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],
                                carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'])
                            current_row += 1
                            current_col = 2

                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
                            current_col += 1
                            for i in index[1::]:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                current_col += 1
                            current_col = 2
                            current_row += 1

                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
                            current_col += 1
                            for i in index[1::]:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                current_col += 1
                            current_col = 2
                            current_row += 1

                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_backoff']
                            current_col += 1
                            for i in index[1::]:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                current_col += 1
                            current_col = 2
                            current_row += 1
                            #4 seperate carriers
                            if bw1+bw2+bw3+bw4 < full_bw - 5:
                                freq_gap = (full_bw - bw1-bw2-bw3-bw4)/3
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw4 - freq_gap - bw3/2))
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw4/2)
                                current_col = 2
                                current_row += 1

                            # 4 close carriers
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1/2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1+bw2/2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1+bw2 + bw3/2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4)/2+bw1+bw2+bw3 + bw4/2)
                            current_col = 2
                            current_row += 1

                            current_row += 2

        # 2 narrow band 2 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup_narrow)):
                bw2 = carrier_setup_narrow[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        index = [i1, i2, i3, i4]
                        if i2 >= i1 and i3 >= i2 and i4 >= i3 and bw1 + bw2 + bw3 + bw4 <= full_bw:
                            worksheet_out.cell(row=current_row, column=1).value = '*'
                            worksheet_out.cell(row=current_row,
                                               column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                carrier_setup_narrow[i1]['carrier_name'], carrier_setup_narrow[i2]['carrier_name'],
                                carrier_setup[i3]['carrier_name'], carrier_setup[i4]['carrier_name'],
                                carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup_narrow[i2]['carrier_testmod'],
                                carrier_setup[i3]['carrier_testmod'], carrier_setup[i4]['carrier_testmod'],
                                carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup_narrow[i2]['carrier_backoff'],
                                carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'])
                            current_row += 1
                            current_col = 2

                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                                'carrier_name']
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2][
                                'carrier_name']
                            current_col += 1
                            for i in index[2::]:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i][
                                    'carrier_name']
                                current_col += 1
                            current_col = 2
                            current_row += 1

                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                                'carrier_testmod']
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2][
                                'carrier_testmod']
                            current_col += 1
                            for i in index[2::]:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i][
                                    'carrier_testmod']
                                current_col += 1
                            current_col = 2
                            current_row += 1

                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1][
                                'carrier_backoff']
                            worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2][
                                'carrier_backoff']
                            current_col += 1
                            for i in index[2::]:
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i][
                                    'carrier_backoff']
                                current_col += 1
                            current_col = 2
                            current_row += 1
                            # 4 seperate carriers
                            if bw1 + bw2 + bw3 + bw4 < full_bw - 5:
                                freq_gap = (full_bw - bw1 - bw2 - bw3 - bw4) / 2
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(
                                    freq_l + bw1 / 2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(
                                    freq_l + bw1 + bw2 / 2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(
                                    freq_l + bw1 + bw2 + freq_gap)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(
                                    freq_h - bw4 / 2)
                                current_col = 2
                                current_row += 1

                            # 4 close carriers
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(
                                freq_m - (bw1 + bw2 + bw3 + bw4) / 2 + bw1 / 2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(
                                freq_m - (bw1 + bw2 + bw3 + bw4) / 2 + bw1 + bw2 / 2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(
                                freq_m - (bw1 + bw2 + bw3 + bw4) / 2 + bw1 + bw2 + bw3 / 2)
                            current_col += 1
                            worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(
                                freq_m - (bw1 + bw2 + bw3 + bw4) / 2 + bw1 + bw2 + bw3 + bw4 / 2)
                            current_col = 2
                            current_row += 1

                            current_row += 2
        self.workbook.save(self.config_file_add)
        # 5 carriers
        worksheet_out = self.workbook.create_sheet("5 Carrier")
        current_row = 1
        # 5 wide band
        for i1 in range(len(carrier_setup)):
            bw1 = carrier_setup[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        for i5 in range(len(carrier_setup)):
                            bw5 = carrier_setup[i5]['carrier_bw']
                            index=[i1, i2, i3, i4, i5]
                            if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and bw1 + bw2 + bw3 + bw4 + bw5 <= full_bw:
                                worksheet_out.cell(row=current_row, column=1).value = '*'
                                worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                    carrier_setup[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],carrier_setup[i5]['carrier_name'],
                                    carrier_setup[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],carrier_setup[i5]['carrier_testmod'],
                                    carrier_setup[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'],carrier_setup[i5]['carrier_backoff'])
                                current_row += 1
                                current_col = 2
                                for i in index:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                for i in index:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                for i in index:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                #5 seperate carriers
                                if bw1+bw2+bw3+bw4+bw5 < full_bw - 5:
                                    freq_gap = (full_bw - bw1-bw2-bw3-bw4)/4
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2 + freq_gap + bw3/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw5 - freq_gap - bw4/2))
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw5/2)
                                    current_col = 2
                                    current_row += 1

                                # 4 close carriers
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2 + bw3/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2+bw3 + bw4/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2+bw3+bw4+bw5/2)
                                current_col = 2
                                current_row += 1

                                current_row += 2

        # 1 narrow band 4 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        for i5 in range(len(carrier_setup)):
                            bw5 = carrier_setup[i5]['carrier_bw']
                            index=[i1, i2, i3, i4, i5]
                            if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and bw1 + bw2 + bw3 + bw4 + bw5 <= full_bw:
                                worksheet_out.cell(row=current_row, column=1).value = '*'
                                worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                    carrier_setup_narrow[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],carrier_setup[i5]['carrier_name'],
                                    carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],carrier_setup[i5]['carrier_testmod'],
                                    carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'],carrier_setup[i5]['carrier_backoff'])
                                current_row += 1
                                current_col = 2

                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
                                current_col += 1
                                for i in index[1::]:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
                                current_col += 1
                                for i in index[1::]:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_backoff']
                                current_col += 1
                                for i in index[1::]:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                #5 seperate carriers
                                if bw1+bw2+bw3+bw4+bw5 < full_bw - 5:
                                    freq_gap = (full_bw - bw1-bw2-bw3-bw4)/4
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2 + freq_gap + bw3/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw5 - freq_gap - bw4/2))
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw5/2)
                                    current_col = 2
                                    current_row += 1

                                # 4 close carriers
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2 + bw3/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2+bw3 + bw4/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2+bw3+bw4+bw5/2)
                                current_col = 2
                                current_row += 1
                                current_row += 2

        # 2 narrow band 3 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup_narrow)):
                bw2 = carrier_setup_narrow[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        for i5 in range(len(carrier_setup)):
                            bw5 = carrier_setup[i5]['carrier_bw']
                            index=[i1, i2, i3, i4, i5]
                            if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and bw1 + bw2 + bw3 + bw4 + bw5 <= full_bw:
                                worksheet_out.cell(row=current_row, column=1).value = '*'
                                worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                    carrier_setup_narrow[i1]['carrier_name'], carrier_setup_narrow[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],carrier_setup[i5]['carrier_name'],
                                    carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup_narrow[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],carrier_setup[i5]['carrier_testmod'],
                                    carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup_narrow[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'],carrier_setup[i5]['carrier_backoff'])
                                current_row += 1
                                current_col = 2

                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_name']
                                current_col += 1
                                for i in index[2::]:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_testmod']
                                current_col += 1
                                for i in index[2::]:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_backoff']
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_backoff']
                                current_col += 1
                                for i in index[2::]:
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                    current_col += 1
                                current_col = 2
                                current_row += 1
                                #5 seperate carriers
                                if bw1+bw2+bw3+bw4+bw5 < full_bw - 5:
                                    freq_gap = (full_bw - bw1-bw2-bw3-bw4)/3
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + bw2/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + bw2 + freq_gap + bw3/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw5 - freq_gap - bw4/2))
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw5/2)
                                    current_col = 2
                                    current_row += 1

                                # 5 close carriers
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2 + bw3/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2+bw3 + bw4/2)
                                current_col += 1
                                worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5)/2+bw1+bw2+bw3+bw4+bw5/2)
                                current_col = 2
                                current_row += 1
                                current_row += 2
        self.workbook.save(self.config_file_add)
        
        # 6 carriers
        worksheet_out = self.workbook.create_sheet("6 Carrier")
        current_row = 1
        # 6 wide band
        for i1 in range(len(carrier_setup)):
            bw1 = carrier_setup[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        for i5 in range(len(carrier_setup)):
                            bw5 = carrier_setup[i5]['carrier_bw']
                            for i6 in range(len(carrier_setup)):
                                bw6 = carrier_setup[i6]['carrier_bw']
                                index=[i1, i2, i3, i4, i5, i6]
                                if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and i6>=i5 and bw1 + bw2 + bw3 + bw4 + bw5 + bw6 <= full_bw:
                                    worksheet_out.cell(row=current_row, column=1).value = '*'
                                    worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                        carrier_setup[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],carrier_setup[i5]['carrier_name'],carrier_setup[i6]['carrier_name'],
                                        carrier_setup[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],carrier_setup[i5]['carrier_testmod'],carrier_setup[i6]['carrier_testmod'],
                                        carrier_setup[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'],carrier_setup[i5]['carrier_backoff'],carrier_setup[i6]['carrier_backoff'])
                                    current_row += 1
                                    current_col = 2
                                    for i in index:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    for i in index:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    for i in index:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    #6 seperate carriers
                                    if bw1+bw2+bw3+bw4+bw5+bw6 < full_bw - 5:
                                        freq_gap = (full_bw - bw1-bw2-bw3-bw4)/5
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2/2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2 + freq_gap + bw3/2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw6 - freq_gap - bw5 - freq_gap - bw4/2))
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw6 - freq_gap - bw5/2))
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw6/2)
                                        current_col = 2
                                        current_row += 1

                                    # 6 close carriers
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2 + bw3/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3 + bw4/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3+bw4+bw5/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3+bw4+bw5+bw6/2)
                                    current_col = 2
                                    current_row += 1

                                    current_row += 2

        # 1 narrow band 5 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        for i5 in range(len(carrier_setup)):
                            bw5 = carrier_setup[i5]['carrier_bw']
                            for i6 in range(len(carrier_setup)):
                                bw6 = carrier_setup[i6]['carrier_bw']
                                index=[i1, i2, i3, i4, i5, i6]
                                if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and i6>=i5 and bw1 + bw2 + bw3 + bw4 + bw5 + bw6 <= full_bw:
                                    worksheet_out.cell(row=current_row, column=1).value = '*'
                                    worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                        carrier_setup_narrow[i1]['carrier_name'], carrier_setup[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],carrier_setup[i5]['carrier_name'],carrier_setup[i6]['carrier_name'],
                                        carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],carrier_setup[i5]['carrier_testmod'],carrier_setup[i6]['carrier_testmod'],
                                        carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'],carrier_setup[i5]['carrier_backoff'],carrier_setup[i6]['carrier_backoff'])
                                    current_row += 1
                                    current_col = 2
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
                                    current_col += 1
                                    for i in index[1::]:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
                                    current_col += 1
                                    for i in index[1::]:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_backoff']
                                    current_col += 1
                                    for i in index[1::]:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    #6 seperate carriers
                                    if bw1+bw2+bw3+bw4+bw5+bw6 < full_bw - 5:
                                        freq_gap = (full_bw - bw1-bw2-bw3-bw4)/5
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2/2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + freq_gap + bw2 + freq_gap + bw3/2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw6 - freq_gap - bw5 - freq_gap - bw4/2))
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw6 - freq_gap - bw5/2))
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw6/2)
                                        current_col = 2
                                        current_row += 1

                                    # 6 close carriers
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2 + bw3/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3 + bw4/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3+bw4+bw5/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3+bw4+bw5+bw6/2)
                                    current_col = 2
                                    current_row += 1

                                    current_row += 2

        # 1 narrow band 5 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup_narrow)):
                bw2 = carrier_setup_narrow[i2]['carrier_bw']
                for i3 in range(len(carrier_setup)):
                    bw3 = carrier_setup[i3]['carrier_bw']
                    for i4 in range(len(carrier_setup)):
                        bw4 = carrier_setup[i4]['carrier_bw']
                        for i5 in range(len(carrier_setup)):
                            bw5 = carrier_setup[i5]['carrier_bw']
                            for i6 in range(len(carrier_setup)):
                                bw6 = carrier_setup[i6]['carrier_bw']
                                index=[i1, i2, i3, i4, i5, i6]
                                if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and i6>=i5 and bw1 + bw2 + bw3 + bw4 + bw5 + bw6 <= full_bw:
                                    worksheet_out.cell(row=current_row, column=1).value = '*'
                                    worksheet_out.cell(row=current_row, column=2).value = '{} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {} {}'.format(
                                        carrier_setup_narrow[i1]['carrier_name'], carrier_setup_narrow[i2]['carrier_name'],carrier_setup[i3]['carrier_name'],carrier_setup[i4]['carrier_name'],carrier_setup[i5]['carrier_name'],carrier_setup[i6]['carrier_name'],
                                        carrier_setup_narrow[i1]['carrier_testmod'], carrier_setup_narrow[i2]['carrier_testmod'],carrier_setup[i3]['carrier_testmod'],carrier_setup[i4]['carrier_testmod'],carrier_setup[i5]['carrier_testmod'],carrier_setup[i6]['carrier_testmod'],
                                        carrier_setup_narrow[i1]['carrier_backoff'], carrier_setup_narrow[i2]['carrier_backoff'], carrier_setup[i3]['carrier_backoff'], carrier_setup[i4]['carrier_backoff'],carrier_setup[i5]['carrier_backoff'],carrier_setup[i6]['carrier_backoff'])
                                    current_row += 1
                                    current_col = 2
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_name']
                                    current_col += 1
                                    for i in index[2::]:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_name']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_testmod']
                                    current_col += 1
                                    for i in index[2::]:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_testmod']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_backoff']
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup_narrow[i2]['carrier_backoff']
                                    current_col += 1
                                    for i in index[2::]:
                                        worksheet_out.cell(row=current_row, column=current_col).value = carrier_setup[i]['carrier_backoff']
                                        current_col += 1
                                    current_col = 2
                                    current_row += 1
                                    #6 seperate carriers
                                    if bw1+bw2+bw3+bw4+bw5+bw6 < full_bw - 5:
                                        freq_gap = (full_bw - bw1-bw2-bw3-bw4)/4
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 / 2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + bw2/2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_l + bw1 + bw2 + freq_gap + bw3/2)
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw6 - freq_gap - bw5 - freq_gap - bw4/2))
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format((freq_h - bw6 - freq_gap - bw5/2))
                                        current_col += 1
                                        worksheet_out.cell(row=current_row, column=current_col).value = '{:.2f}M'.format(freq_h - bw6/2)
                                        current_col = 2
                                        current_row += 1

                                    # 6 close carriers
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2 + bw3/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3 + bw4/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3+bw4+bw5/2)
                                    current_col += 1
                                    worksheet_out.cell(row=current_row, column=current_col).value = '{}M'.format(freq_m-(bw1+bw2+bw3+bw4+bw5+bw6)/2+bw1+bw2+bw3+bw4+bw5+bw6/2)
                                    current_col = 2
                                    current_row += 1

                                    current_row += 2

        self.workbook.save(self.config_file_add)


if __name__ == '__main__':
    tx_rf_func = TxRfFuncConfig('D:\RuPyTest_P1A03\OperateString\TestDefGeneration\TX_RF_FUNC\TxRfFunc_Config_V1.xlsx')
    tx_rf_func.gen_carrier_combination()
    #tx_rf_func.gen_carriers_TxRfFunc('6 Carrier')