from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import copy
from decimal import Decimal
import re
import time

class TxRfFuncConfig:
    def __init__(self, config_file_add):
        self.config_file_add = config_file_add
        self.workbook = load_workbook(config_file_add)
        backup_filename =config_file_add + '_backup'
        self.workbook.save(backup_filename)
        #self.workbook.active
        self.sheets = self.workbook.sheetnames
        self.load_test_config()
        self.load_common_config()
        self.load_carrier_limit_table()
        self.carrier_sheets = self.workbook.sheetnames[4::]
        self.max_power_total = Decimal(self.workbook['CarrierSetup'].cell(row=6, column=2).value)


    def load_test_config(self):
        worksheet = self.workbook['TestConfig']
        self.testDef_head = worksheet.cell(row=2, column=2).value
        self.output_file_name = worksheet.cell(row=3, column=2).value
        self.start_index = int(worksheet.cell(row=4, column=2).value)
        self.trdc_delay_short   = worksheet.cell(row=5, column=2).value
        self.trdc_delay_long    = worksheet.cell(row=6, column=2).value
        self.branch = worksheet.cell(row=7, column=2).value
        self.AXC = worksheet.cell(row=8, column=2).value

    def load_common_config(self):
        worksheet = self.workbook['CommonConfig']
        max_row = worksheet.max_row
        self.para_tap = []
        self.para_value = []
        for i in range(2, max_row+1):
            self.para_tap.append(worksheet.cell(row=i, column=1).value)
            self.para_value.append(worksheet.cell(row=i, column=2).value)
        self.para_value_txblock = copy.deepcopy(self.para_value)
        for i in range(len(self.para_value_txblock)):
            self.para_value_txblock[i] = '-'
        index = self.para_tap.index('delayAfterTxOn')
        self.para_value_txblock[index] = 0
        index = self.para_tap.index('doMeas')
        self.para_value_txblock[index] = 0
        index = self.para_tap.index('measTime')
        self.para_value_txblock[index] = 0

    def load_carrier_limit_table(self):
        worksheet = self.workbook['CarrierLimit']
        max_row = worksheet.max_row
        self.carrier_limit_table = []
        for i in range(2,max_row+1):
            self.carrier_limit_table.append(list(worksheet.iter_rows(i, i, worksheet.min_column, worksheet.max_column, True))[0])

    def load_carrier_config(self, worksheet_name):
        worksheet = self.workbook[worksheet_name]
        max_row = worksheet.max_row
        current_row = 1
        carriers_config = []
        while current_row < max_row:
            row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            current_row = current_row + 1
            '''
            All the carrier configs must start with * at the first col
            '''
            if row_data[0] == r'*':
                print(row_data[1])
                row_data = list(filter(None, row_data))
                carrier_dict={}
                carrier_dict['commont'] = row_data[1]
                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                row_data = list(filter(None, row_data))
                current_row = current_row + 1
                carrier_dict['carrier_type'] = row_data

                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                row_data = list(filter(None, row_data))
                current_row = current_row + 1
                carrier_dict['test_mod'] = row_data

                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                row_data = list(filter(None, row_data))
                current_row = current_row + 1
                carrier_dict['power_backoff'] = row_data

                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]

                carrier_dict['freq'] = []
                while(row_data[0]==None and row_data[1]):
                    row_data = list(filter(None, row_data))
                    carrier_dict['freq'].append(row_data)
                    current_row = current_row + 1
                    row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column,True))[0]
                carriers_config.append(carrier_dict)
        return carriers_config

    def gen_carrier_para(self, carrier_dict):
        #if or not test OBUE, EVM, carrier power and test limit based on carrier combination
        para_value_carrier = copy.deepcopy(self.para_value)
        spurEm=0
        evm=0
        carrPwr=0
        delayAfterTxOn=self.trdc_delay_short
        pwrAcc=99
        evmMax=99
        evmMaxIOT=99
        RBW='1000k'
        limitTable='BC1_NR2'

        if len(carrier_dict['carrier_type']) <= 2:
            for i in range(len(carrier_dict['carrier_type'])):
                if carrier_dict['carrier_type'][i] != 'L4':
                    spurEm = 0
                    carrPwr = 0
                    evm = 0
                    break
                else:
                    spurEm = 1
                    carrPwr = 1
                    evm = 1
        for i in range(len(carrier_dict['carrier_type'])):
            for config in self.carrier_limit_table:
                if carrier_dict['carrier_type'][i] == config[0] and carrier_dict['test_mod'][i] == config[1] and carrier_dict['power_backoff'][i] == config[2]:
                    if config[2] == '0dB' or config[2] == '-0dB':
                        spurEm = 1
                        delayAfterTxOn = self.trdc_delay_long

                    if len(carrier_dict['carrier_type']) ==1:
                        if config[3] != 'NA':
                            if config[0] == 'L4':
                                evmMaxIOT = min(evmMaxIOT, float(config[3]))
                            else:
                                evm = 1
                            evmMax = min(evmMax, float(config[3]))
                        if config[5] != 'NA':
                            carrPwr = 1
                            pwrAcc = min(pwrAcc, float(config[5]))
                    else:
                        if config[4] != 'NA':
                            if config[0] == 'L4':
                                evmMaxIOT = min(evmMaxIOT, float(config[4]))
                            else:
                                evm = 1
                            evmMax = min(evmMax, float(config[4]))
                        if config[6] != 'NA' and config[0] != 'L4':
                            carrPwr = 1
                            pwrAcc = min(pwrAcc, float(config[6]))
                    RBW = '{}k'.format(min(int(RBW.strip('k')), int(config[7].strip('k'))))
                    if config[8] != 'BC1_NR2':
                        limitTable = config[8]

        index = self.para_tap.index('spurEm')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = spurEm

        index = self.para_tap.index('evm')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = evm

        index = self.para_tap.index('carrPwr')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = carrPwr

        index = self.para_tap.index('delayAfterTxOn')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = delayAfterTxOn

        index = self.para_tap.index('pwrAcc')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = pwrAcc

        index = self.para_tap.index('evmMax')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = evmMax

        index = self.para_tap.index('evmMaxIOT')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = evmMaxIOT

        index = self.para_tap.index('RBW')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = RBW

        index = self.para_tap.index('limitTable')
        if self.para_value[index] == 'NA':
            para_value_carrier[index] = limitTable
        return para_value_carrier

    def gen_carriers_TxRfFunc(self, worksheet_name):
        self.start_index = 1
        output_file_name = '{}-{}.txt'.format(self.output_file_name.strip('.txt'), worksheet_name)
        with open(output_file_name, 'w') as file_out:
            with open(self.testDef_head, 'r') as finle_in:
                testDef_head = finle_in.readlines()
            file_out.writelines(testDef_head)
            file_out.write('\n')

        worksheet = self.workbook[worksheet_name]
        max_row = worksheet.max_row
        current_row = 1
        while current_row < max_row:
            row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            current_row = current_row + 1
            '''
            All the carrier configs must start with * at the first col
            '''
            if row_data[0] == r'*':
                print(row_data[1])
                row_data = list(filter(None, row_data))
                carrier_dict={}
                carrier_dict['commont'] = row_data[1]
                carrier_name = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                carrier_name = list(filter(None, carrier_name))
                current_row = current_row + 1
                carrier_dict['carrier_type'] = carrier_name

                carrier_testmod = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                carrier_testmod = list(filter(None, carrier_testmod))
                current_row = current_row + 1
                carrier_dict['test_mod'] = carrier_testmod



                carrier_power = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                carrier_power = list(filter(None, carrier_power))
                current_row = current_row + 1
                for i in range(len(carrier_power)):
                    if carrier_name[i] == 'L4' and self.max_power_total/(len(carrier_power))*pow(10, Decimal(carrier_power[i].strip('dB')))>20:
                        carrier_power[i] = '20W'

                carrier_dict['power_backoff'] = carrier_power

                carrier_freq = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                carrier_dict['freq'] = []
                while(carrier_freq[0]==None and carrier_freq[1]):
                    carrier_freq = list(filter(None, carrier_freq))
                    carrier_dict['freq'].append(carrier_freq)
                    current_row = current_row + 1
                    carrier_freq = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column,True))[0]

                carrier_config = carrier_dict

                print('Generating TX_RF_FUNC_{}'.format(self.start_index))
                with open(output_file_name, 'a') as f:
                    f.write('\n*{}\n'.format(carrier_config['commont']))
                    f.write('#TX_RF_FUNC_{:<10}chConf'.format(self.start_index, ''))
                    f.write('{:<360}'.format(''))
                    for i in self.para_tap:
                        f.write('{:<20}'.format(i))
                    f.write('\n*----------------------------------------------------------------------------------------------------------------------------------\n')
                    para_value_carrier = self.gen_carrier_para(carrier_config)
                    for freq in carrier_config['freq']:
                        f.write('{:<28}'.format(''))
                        carrier_setup_string = ''
                        for i in range(len(freq)):
                            carrier_setup_string += '<{}_{}_{}_Dd_MOD{}_P{}_PROTtrdcsetup7_AXC{}bit>;'.format(self.branch,carrier_config['carrier_type'][i], freq[i],carrier_config['test_mod'][i],carrier_config['power_backoff'][i], self.AXC)
                        f.write('{:<360} '.format(carrier_setup_string))


                        for i in para_value_carrier:
                            f.write('{:<20}'.format(i))
                        f.write('\n')
                        f.write('{:<389}'.format('<TRblock> '))
                        for i in self.para_value_txblock:
                            f.write('{:<20}'.format(i))
                        f.write('\n')

                print(self.start_index)
                self.start_index = self.start_index + 1
        with open(output_file_name, 'r') as f:
            data = f.readlines()
            index_range = '{}:{}'.format(1,self.start_index - 1 )
            for i in range(len(data)):
                if re.search('RunIndexRange', data[i]):
                    data[i] = re.sub('RunIndexRange',index_range ,data[i])

        with open(output_file_name, 'w') as f:
            f.writelines(data)

    def check_carrier_combin_availible(self, carrier_comb):
        type = []
        bandwidth = []
        testmod = []
        backoff = []
        is_narrow_band = []
        num_of_FB = []
        for carrier in carrier_comb:
            type.append(carrier['carrier_name'])
            bandwidth.append(carrier['carrier_bw'] )
            testmod.append(carrier['carrier_testmod'] )
            backoff.append(carrier['carrier_backoff'])
            num_of_FB.append(carrier['carrier_FB'])
        #NB, W and NR cannot set up together
        if sum(bandwidth) > self.full_bw:
            return False
        if sum(num_of_FB) > self.max_num_carrier:
            return False
        if {'L4', 'W'}.issubset(type):
            for t in type:
                if re.match('NR', t):
                    return False
        return True

    def correct_backoff_power(self,carrier_comb_narrow, carrier_comb):
        for i in range(len(carrier_comb_narrow)):
            carrier_comb_narrow[i]['carrier_backoff'] = '{}dB'.format(carrier_comb_narrow[i]['carrier_backoff'])
        for i in range(len(carrier_comb)):
            carrier_comb[i]['carrier_backoff'] = '{}dB'.format(carrier_comb[i]['carrier_backoff'])
        return carrier_comb_narrow,carrier_comb

    def gen_carrier_freq(self, worksheet_out, carrier_comb_narrow, carrier_comb):
        bw = []
        carrier_name = []
        carrier_testmod = []
        carrier_backoff = []
        comment = ''
        carrier_comb_narrow, carrier_comb = self.correct_backoff_power(carrier_comb_narrow, carrier_comb)

        carrier_comb = carrier_comb_narrow + carrier_comb
        for carrier in carrier_comb:
            bw.append(carrier['carrier_bw'])
            carrier_name.append(carrier['carrier_name'])
            carrier_testmod.append(carrier['carrier_testmod'])
            carrier_backoff.append(carrier['carrier_backoff'])

        for i in range(len(carrier_comb)):
            comment += '{} '.format(carrier_name[i])
        for i in range(len(carrier_comb)):
            comment += '{} '.format(carrier_testmod[i])
        for i in range(len(carrier_comb)):
            comment += '{} '.format(carrier_backoff[i])

        worksheet_out.cell(row=self.current_row, column=1).value = '*'
        worksheet_out.cell(row=self.current_row, column=2).value = comment
        self.current_row +=1
        self.current_col = 2
        for carrier in carrier_comb:
            worksheet_out.cell(row=self.current_row, column=self.current_col).value = carrier['carrier_name']
            self.current_col +=1
        self.current_row += 1
        self.current_col = 2
        for carrier in carrier_comb:
            worksheet_out.cell(row=self.current_row, column=self.current_col).value = carrier['carrier_testmod']
            self.current_col +=1
        self.current_row += 1
        self.current_col = 2
        for carrier in carrier_comb:
            worksheet_out.cell(row=self.current_row, column=self.current_col).value = carrier['carrier_backoff']
            self.current_col +=1
        self.current_row += 1


        if self.close_carrier=='True' and sum(bw) + self.close_gap * (len(carrier_comb) - 1) < self.full_bw:
            carrier_gap = self.close_gap
            if len(carrier_comb_narrow)==2 and carrier_gap + bw[0] + bw[1] > 20:
                carrier_gap = 20- bw[0] - bw[1]

            #L band enable
            if self.workbook['CarrierSetup'].cell(row=7, column=2).value == 'True':
                bw_temp = 0
                self.current_col = 2
                for i in range(len(carrier_comb)):
                    worksheet_out.cell(row=self.current_row, column=self.current_col).value = '{:.2f}M'.format(self.freq_l + bw[i]/2 + bw_temp)
                    bw_temp += bw[i] + carrier_gap
                    self.current_col += 1
                self.current_row += 1

            #H band enable
            if self.workbook['CarrierSetup'].cell(row=8, column=2).value == 'True':
                bw_temp = 0
                self.current_col = 2
                for i in range(len(carrier_comb)):
                    worksheet_out.cell(row=self.current_row, column=self.current_col).value = '{:.2f}M'.format(self.freq_h - bw[i]/2 - bw_temp)
                    bw_temp += bw[i] + carrier_gap
                    self.current_col += 1
                self.current_row += 1

            #M band enable
            if self.workbook['CarrierSetup'].cell(row=9, column=2).value == 'True':
                bw_temp = 0
                bw_delta = sum(bw) + carrier_gap * (len(carrier_comb) - 1)
                self.current_col = 2
                for i in range(len(carrier_comb)):
                    worksheet_out.cell(row=self.current_row, column=self.current_col).value = '{:.2f}M'.format(self.freq_m - bw_delta/2 + bw[i]/2 + bw_temp)
                    bw_temp += bw[i] + carrier_gap
                    self.current_col += 1
                self.current_row += 1


        if self.seperate_carrier =='True':
            carrier_gap = (self.full_bw - sum(bw))/(len(carrier_comb)-1)
            if len(carrier_comb_narrow)==2 and carrier_gap + bw[0] + bw[1] > 20:
                carrier_gap = 20- bw[0] - bw[1]
            bw_temp = 0
            bw_delta = sum(bw) + carrier_gap * (len(carrier_comb) - 1)
            self.current_col = 2
            for i in range(len(carrier_comb)):
                worksheet_out.cell(row=self.current_row, column=self.current_col).value = '{:.2f}M'.format(self.freq_m - bw_delta/2 + bw[i]/2 + bw_temp)
                bw_temp += bw[i] + carrier_gap
                self.current_col += 1
        self.current_row += 2
        self.current_col = 2

    def gen_carrier_combination(self):
        worksheet_in = self.workbook['CarrierSetup']

        max_row = worksheet_in.max_row
        self.freq_l = Decimal(worksheet_in.cell(row=1, column=2).value)
        self.freq_h = Decimal(worksheet_in.cell(row=2, column=2).value)
        self.max_num_carrier = int(worksheet_in.cell(row=3, column=2).value)
        self.freq_m = (self.freq_l + self.freq_h)/2
        self.full_bw = self.freq_h - self.freq_l
        self.close_carrier = worksheet_in.cell(row=4, column=2).value
        self.seperate_carrier = worksheet_in.cell(row=10, column=2).value
        self.close_gap = Decimal(worksheet_in.cell(row=5, column=2).value)
        self.max_power_total = Decimal(worksheet_in.cell(row=6, column=2).value)


        offset = 20
        carrier_setup = []
        carrier_setup_narrow = []


        for row in range(offset, max_row+1):
            carrier_dict={}
            carrier_dict['carrier_name'] = worksheet_in.cell(row=row, column=1).value
            carrier_dict['carrier_bw'] = Decimal(worksheet_in.cell(row=row, column=2).value)
            carrier_dict['carrier_testmod'] = worksheet_in.cell(row=row, column=3).value
            carrier_dict['carrier_backoff'] = Decimal(worksheet_in.cell(row=row, column=4).value)
            carrier_dict['carrier_FB'] = int(worksheet_in.cell(row=row, column=6).value)
            if worksheet_in.cell(row=row, column=5).value =='1':
                carrier_setup_narrow.append(carrier_dict)
            else:
                carrier_setup.append(carrier_dict)

        worksheet_out = self.workbook.create_sheet("tm3p1 0dB")
        self.current_row = 1
        self.current_col = 1

        #Single carrier
        # Single carrier wide band
        for i1 in range(len(carrier_setup)):
            bw = carrier_setup[i1]['carrier_bw']
            worksheet_out.cell(row=self.current_row, column=1).value = '*'
            worksheet_out.cell(row=self.current_row, column=2).value = '{} {} {}dB'.format(carrier_setup[i1]['carrier_name'],carrier_setup[i1]['carrier_testmod'],carrier_setup[i1]['carrier_backoff'] )
            self.current_row += 1
            current_col = 2
            worksheet_out.cell(row=self.current_row, column=current_col).value = carrier_setup[i1]['carrier_name']
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = carrier_setup[i1]['carrier_testmod']
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}dB'.format(carrier_setup[i1]['carrier_backoff'])
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}M'.format(self.freq_l+bw/2)
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}M'.format(self.freq_m)
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}M'.format(self.freq_h-bw/2)
            self.current_row += 2
        # Single carrier narrow band
        for i1 in range(len(carrier_setup_narrow)):
            bw = carrier_setup_narrow[i1]['carrier_bw']
            worksheet_out.cell(row=self.current_row, column=1).value = '*'
            worksheet_out.cell(row=self.current_row, column=2).value = '{} {} {}dB'.format(carrier_setup_narrow[i1]['carrier_name'],carrier_setup_narrow[i1]['carrier_testmod'],carrier_setup_narrow[i1]['carrier_backoff'])
            self.current_row += 1
            current_col = 2
            worksheet_out.cell(row=self.current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_name']
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = carrier_setup_narrow[i1]['carrier_testmod']
            self.current_row += 1
            p = self.max_power_total*pow(10, carrier_setup_narrow[i1]['carrier_backoff']/10)
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}dB'.format(carrier_setup_narrow[i1]['carrier_backoff'])
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}M'.format(self.freq_l + bw / 2)
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}M'.format(self.freq_m)
            self.current_row += 1
            worksheet_out.cell(row=self.current_row, column=current_col).value = '{}M'.format(self.freq_h - bw / 2)
            self.current_row += 2
        self.workbook.save(self.config_file_add)

        # 2 carriers
        # 2 wide band
        self.current_col = 2
        for i1 in range(len(carrier_setup)):
            for i2 in range(len(carrier_setup)):
                carrier_comb_narrow = []
                carrier_comb = [copy.deepcopy(carrier_setup[i1]),copy.deepcopy(carrier_setup[i2])]
                if i2 >= i1 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                    self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)
        # 1 narrow band 1 wide band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup)):
                bw2 = carrier_setup[i2]['carrier_bw']
                carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1])]
                carrier_comb = [copy.deepcopy(carrier_setup[i2])]
                if i2 >= i1 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                    self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)

        # 2 narrow band
        for i1 in range(len(carrier_setup_narrow)):
            bw1 = carrier_setup_narrow[i1]['carrier_bw']
            for i2 in range(len(carrier_setup_narrow)):
                bw2 = carrier_setup_narrow[i2]['carrier_bw']
                carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1]), copy.deepcopy(carrier_setup_narrow[i2])]
                carrier_comb = []
                if i2 >= i1 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                    self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)

        self.workbook.save(self.config_file_add)

        # 3 carriers
        self.current_col = 2
        # 3 wide band
        for i1 in range(len(carrier_setup)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    carrier_comb_narrow = []
                    carrier_comb = [copy.deepcopy(carrier_setup[i1]), copy.deepcopy(carrier_setup[i2]),copy.deepcopy(carrier_setup[i3])]
                    if i2 >= i1 and i3 >=i2 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                        self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)
        # 1 narrow band 2 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1])]
                    carrier_comb = [copy.deepcopy(carrier_setup[i2]),copy.deepcopy(carrier_setup[i3])]
                    if i3 >=i2 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                        self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)

        # 2 narrow band 1 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup_narrow)):
                for i3 in range(len(carrier_setup)):
                    carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1]),copy.deepcopy(carrier_setup_narrow[i2])]
                    carrier_comb = [copy.deepcopy(carrier_setup[i3])]
                    if i2 >= i1 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                        self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)


        # 4 carriers
        self.current_col = 2
        # 4 wide band
        for i1 in range(len(carrier_setup)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        carrier_comb_narrow = []
                        carrier_comb = [copy.deepcopy(carrier_setup[i1]), copy.deepcopy(carrier_setup[i2]),copy.deepcopy(carrier_setup[i3]),copy.deepcopy(carrier_setup[i4])]
                        if i2 >= i1 and i3 >=i2 and i4>=i3 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                            self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)


        # 1 narrow band 3 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1])]
                        carrier_comb = [copy.deepcopy(carrier_setup[i2]), copy.deepcopy(carrier_setup[i3]), copy.deepcopy(carrier_setup[i4])]
                        if i3 >=i2 and i4>=i3 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                            self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)
        # 2 narrow band 2 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup_narrow)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1]),
                                               copy.deepcopy(carrier_setup_narrow[i2])]
                        carrier_comb = [copy.deepcopy(carrier_setup[i3]),copy.deepcopy(carrier_setup[i4])]
                        if i2 >= i1 and i4>=i3 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                            self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)

        # 5 carriers
        self.current_col = 2
        # 5 wide band
        for i1 in range(len(carrier_setup)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        for i5 in range(len(carrier_setup)):
                            carrier_comb_narrow = []
                            carrier_comb = [copy.deepcopy(carrier_setup[i1]), copy.deepcopy(carrier_setup[i2]),
                                            copy.deepcopy(carrier_setup[i3]), copy.deepcopy(carrier_setup[i4]), copy.deepcopy(carrier_setup[i5])]
                            if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                                self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)

        # 1 narrow band 4 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        for i5 in range(len(carrier_setup)):
                            carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1])]
                            carrier_comb = [copy.deepcopy(carrier_setup[i2]), copy.deepcopy(carrier_setup[i3]),
                                            copy.deepcopy(carrier_setup[i4]), copy.deepcopy(carrier_setup[i5])]
                            if i3 >=i2 and i4>=i3 and i5>=i4 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                                self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)

        # 2 narrow band 3 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup_narrow)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        for i5 in range(len(carrier_setup)):
                            carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1]),
                                                   copy.deepcopy(carrier_setup_narrow[i2])]
                            carrier_comb = [copy.deepcopy(carrier_setup[i3]), copy.deepcopy(carrier_setup[i4]),
                                            copy.deepcopy(carrier_setup[i5])]
                            if i2 >= i1 and i4>=i3 and i5>=i4 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                                self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)
        
        # 6 carriers

        self.current_col = 2
        # 6 wide band
        for i1 in range(len(carrier_setup)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        for i5 in range(len(carrier_setup)):
                            for i6 in range(len(carrier_setup)):
                                carrier_comb_narrow = []
                                carrier_comb = [copy.deepcopy(carrier_setup[i1]), copy.deepcopy(carrier_setup[i2]),
                                                copy.deepcopy(carrier_setup[i3]), copy.deepcopy(carrier_setup[i4]),
                                                copy.deepcopy(carrier_setup[i5]), copy.deepcopy(carrier_setup[i6])]
                                if i2 >= i1 and i3 >=i2 and i4>=i3 and i5>=i4 and i6>=i5 and self.check_carrier_combin_availible(
                                        carrier_comb_narrow+carrier_comb):
                                    self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)

        # 1 narrow band 5 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        for i5 in range(len(carrier_setup)):
                            for i6 in range(len(carrier_setup)):
                                carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1])]
                                carrier_comb = [copy.deepcopy(carrier_setup[i2]), copy.deepcopy(carrier_setup[i3]),
                                                copy.deepcopy(carrier_setup[i4]), copy.deepcopy(carrier_setup[i5]), copy.deepcopy(carrier_setup[i6])]
                                if i3 >=i2 and i4>=i3 and i5>=i4 and i6>=i5 and self.check_carrier_combin_availible(
                                        carrier_comb_narrow+carrier_comb):
                                    self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)


        # 2 narrow band 5 wide band
        for i1 in range(len(carrier_setup_narrow)):
            for i2 in range(len(carrier_setup_narrow)):
                for i3 in range(len(carrier_setup)):
                    for i4 in range(len(carrier_setup)):
                        for i5 in range(len(carrier_setup)):
                            for i6 in range(len(carrier_setup)):
                                carrier_comb_narrow = [copy.deepcopy(carrier_setup_narrow[i1]),
                                                       copy.deepcopy(carrier_setup_narrow[i2])]
                                carrier_comb = [copy.deepcopy(carrier_setup[i3]), copy.deepcopy(carrier_setup[i4]), copy.deepcopy(carrier_setup[i5]), copy.deepcopy(carrier_setup[i6])]
                                if i2 >= i1 and i4>=i3 and i5>=i4 and i6>=i5 and self.check_carrier_combin_availible(carrier_comb_narrow+carrier_comb):
                                    self.gen_carrier_freq(worksheet_out, carrier_comb_narrow, carrier_comb)
        self.workbook.save(self.config_file_add)
        self.carrier_sheets = self.workbook.sheetnames[4::]
    def replace_0dBtm3p1_testmod_bcckoff(self, test_mod, backoff_power):
        worksheet_in = self.workbook['tm3p1 0dB']
        worksheet_out = self.workbook.create_sheet('{} {}'.format(test_mod, backoff_power))
        max_row = worksheet_in.max_row
        max_col = worksheet_in.max_column
        for i in range(1, max_row + 1):
            for j in range(1, max_col + 1):
                data = worksheet_in.cell(row=i, column=j).value
                if data:
                    data = re.sub('tm3p1', test_mod, data)
                    data = re.sub('0dB', backoff_power, data)
                    worksheet_out.cell(row=i, column=j).value =data
        self.workbook.save(self.config_file_add)
        self.carrier_sheets = self.workbook.sheetnames[4::]

if __name__ == '__main__':
    tx_rf_func: TxRfFuncConfig = TxRfFuncConfig('D:\RuPyTest_P1A03\OperateString\TestDefGeneration\TX_RF_FUNC\Pamukkale_R1B_Regression.xlsx')

    #tx_rf_func.gen_carrier_combination()

    #tx_rf_func.replace_0dBtm3p1_testmod_bcckoff('tm3p1', '-1dB')
    #tx_rf_func.replace_0dBtm3p1_testmod_bcckoff('tm3p1a', '0dB')
    #tx_rf_func.replace_0dBtm3p1_testmod_bcckoff('tm3p1a', '-0.8dB')
    #tx_rf_func.replace_0dBtm3p1_testmod_bcckoff('tm3p1a', '-1dB')

    for worksheet in tx_rf_func.carrier_sheets:
        tx_rf_func.gen_carriers_TxRfFunc(worksheet)


