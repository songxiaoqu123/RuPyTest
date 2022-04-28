from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

input_filename = 'TxRfFunc_PamukkaleR1B_V2.xlsx'
output_filename = 'TxRfFunc_PamukkaleR1B.txt'
#sheet = 'tm3p1 0Backoff 5 Carrier'
sheet = 'NODE'
AXC = 30
start_index = 9100
Branch = 'A'
TRblock = '					  <TRblock> 																											  -			          - 				  - 			      0 				  -   			       -                   -                   -                  -                    -                  0                   0                      -                -                    -                  -                   -                    -                  -                    -                  -                   -                   -                   -                   -                   -                   -                    -                  -                   -                     -\n'

workbook = load_workbook(input_filename)
workbook.active
worksheet = workbook[sheet]

max_row = worksheet.max_row
current_row = 1

'''
get common config tab in the 1st row
get common config data in the 2nd row
get case config tab in the 3rd row

'''
row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
common_tab = list(filter(None,row_data))
current_row = current_row + 1
row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
common_config = list(filter(None, row_data))
current_row = current_row + 1
row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
case_tab = list(filter(None,row_data))
current_row = current_row + 1

print(common_tab)
print(common_config)
print(case_tab)
tab_common_string = '{:<120}'.format('chConf')
for t in common_tab:
    tab_common_string = tab_common_string + '{:<20}'.format(t)
for t in case_tab:
    tab_common_string = tab_common_string + '{:<20}'.format(t)
print(tab_common_string)





with open(output_filename, 'w') as f:
    while current_row < max_row:
        row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
        current_row = current_row + 1
        '''
        if the first cell is *, it will be a comment to be printed in testDef
        elseif the 1st cell of the line exist and not a number, a new test case will be created.
        '''
        if row_data[0] == r'*':
            f.write('*{}\n'.format(row_data[1]))
        elif row_data[0]:
            tab_string = '#TX_RF_FUNC_{:<10}{}\n'.format(start_index, tab_common_string)
            start_index = start_index + 1
            f.write(tab_string)
            f.write('*----------------------------------------------------------------------------------------------------------------------------------\n')
            conf_string=''
            for c in common_config:
                conf_string = conf_string + '{:<20}'.format(c)
            for c in row_data:
                conf_string = conf_string + '{:<20}'.format(c)
            #get carrier type
            row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            carrier_type = list(filter(None, row_data))
            current_row = current_row + 1
            print(carrier_type)

            # get test mode
            row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            test_mode = list(filter(None, row_data))
            current_row = current_row + 1
            print(test_mode)

            # get power backoff
            row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            back_off = list(filter(None, row_data))
            current_row = current_row + 1
            print(back_off)
            row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            while(row_data[0]==None and row_data[1]):
                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
                freq = list(filter(None, row_data))
                print(freq)
                carrier_setup_string = ''
                for i in range(len(carrier_type)):
                    carrier_setup_string = carrier_setup_string + '<{}_{}_{}_Dd_MOD{}_P{}_PROTtrdcsetup7_AXC{}bit>;'.format(Branch,carrier_type[i], freq[i],test_mode[i],back_off[i], AXC)
                carrier_setup_string = carrier_setup_string + ' '
                f.write('{:<22}{:<120}{}\n'.format('',carrier_setup_string, conf_string))
                f.write(TRblock)
                print(conf_string)

                current_row = current_row + 1
                row_data = list(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))[0]
            f.write('\n')
















'''
get test case 
read first 2 row of excel
'''




'''
with open(output_filename, 'w') as f:
    while current_row <= max_row:

            if worksheet.cell(current_row, 1).value:

                data = tuple(worksheet.iter_rows(current_row, current_row, worksheet.min_column, worksheet.max_column, True))
                print(data)
            current_row = current_row + 1
'''