from openpyxl import Workbook

branches = ['A', 'B']
mpa_index = [0, 1]
dpa_index = [0, 1,2,3]
mpa_offset = [[[1717, 1717], [1497, 1497]], [[1595, 1595], [1602, 1601]]]
pa_final_temp = [[['25', '25.3'], ['25', '25.5']], [['27.3', '28'], ['26.8', '26.8']]]
pa_driver_temp = [[[3199, 3199], [3400, 3401], [3301, 3300], [3018, 3005]],
                  [[3200, 3201], [3403, 3416], [3313, 3314], [3075, 3060]]]
dpa_offset = [[['24', '24.5'], ['24.3', '24.5'], ['24.5', '24'], ['24', '24.8']],
              [['24.8', '25'], ['24.5', '24.5'], ['24.5', '25'], ['25', '25']]]
tor_temp = [['27.5', '28'], ['27.5', '28']]
tor_outputpower = [[1.5628, 1.593742], [1.619975, 1.588878]]

rx_temp = [['7.5', '8'], ['7.5', '8']]
rx_highgain_noatt = [[11.5628, 11.593742], [11.619975, 11.588878]]
rx_highgain_att= [[21.5628, 21.593742], [21.619975, 31.588878]]
rx_bypass_noatt= [[31.5628, 31.593742], [31.619975,31.588878]]
rx_bypass_att= [[41.5628, 41.593742], [11.619975, 41.588878]]





wb = Workbook()
wb.remove(wb.active)
##########write PA to excel###############
ws_pa = wb.create_sheet("PA")
dpa_col_offs = len(branches) * len(mpa_index) * 3
for b in range(len(branches)):
    for i in range(len(mpa_index)):
        ws_pa.cell(row=1, column=b * len(mpa_index) * 3 + i * 3 + 1).value = 'MPA {}.{}'.format(b, i)
        ws_pa.cell(row=2, column=b*len(mpa_index) * 3 + i * 3 + 1).value = 'Temp'
        ws_pa.cell(row=2, column=b*len(mpa_index) * 3 + i * 3 + 2).value = 'Offset'
        for j in range(len(pa_final_temp[b][i])):
            ws_pa.cell(row=j + 3, column=b * len(mpa_index) * 3 + i * 3 + 1).value = pa_final_temp[b][i][j]
            ws_pa.cell(row=j + 3, column=b * len(mpa_index) * 3 + i * 3 + 2).value = mpa_offset[b][i][j]

    for i in range(len(dpa_index)):
        ws_pa.cell(row=1, column=dpa_col_offs + b*len(dpa_index) * 3 + i * 3 + 1).value = 'DPA {}.{}'.format(b, i)
        ws_pa.cell(row=2, column=dpa_col_offs + b*len(dpa_index) * 3 + i * 3 + 1).value = 'Temp'
        ws_pa.cell(row=2, column=dpa_col_offs + b*len(dpa_index) * 3 + i * 3 + 2).value = 'Offset'
        for j in range(len(pa_driver_temp[b][i])):
            ws_pa.cell(row=j + 3, column=dpa_col_offs + b * len(dpa_index) * 3 + i * 3 + 1).value = pa_driver_temp[b][i][j]
            ws_pa.cell(row=j + 3, column=dpa_col_offs + b * len(dpa_index) * 3 + i * 3 + 2).value = dpa_offset[b][i][j]


##########write TX to excel###############
ws_tx = wb.create_sheet("Tx")
for b in range(len(branches)):
    ws_tx.cell(row=1, column=b * 3 + 1).value = 'Tx {}'.format(b)
    ws_tx.cell(row=2, column=b * 3 + 1).value = 'Temp'
    ws_tx.cell(row=2, column=b * 3 + 2).value = 'Power'
    for j in range(len(tor_temp[b])):
        ws_tx.cell(row=j + 3, column=b * 3 + 1).value = tor_temp[b][j]
        ws_tx.cell(row=j + 3, column=b * 3 + 2).value = tor_outputpower[b][j]

##########write RX to excel###############
ws_tx = wb.create_sheet("Rx")
for b in range(len(branches)):
    ws_tx.cell(row=1, column=b * 6 + 1).value = 'Rx {}'.format(b)
    ws_tx.cell(row=2, column=b * 6 + 1).value = 'Temp'
    ws_tx.cell(row=2, column=b * 6 + 2).value = 'HighGainNoAtt'
    ws_tx.cell(row=2, column=b * 6 + 3).value = 'HighGainAtt'
    ws_tx.cell(row=2, column=b * 6 + 4).value = 'BypassNoAtt'
    ws_tx.cell(row=2, column=b * 6 + 5).value = 'BypassAtt'
    for j in range(len(rx_temp[b])):
        ws_tx.cell(row=j + 3, column=b * 6 + 1).value = rx_temp[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 2).value = rx_highgain_noatt[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 3).value = rx_highgain_att[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 4).value = rx_bypass_noatt[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 5).value = rx_bypass_att[b][j]



wb.save('outp.xlsx')
