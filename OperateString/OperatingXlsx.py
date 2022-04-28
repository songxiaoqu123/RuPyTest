#coding=utf-8
import sys
import os
import openpyxl

class Excel:
    def __init__(self, file_path):
        self.file_path = file_path

    def load_excel(self):
        '''
        加载excel
        '''
        open_excel = openpyxl.load_workbook(self.file_path)#拿到excel的所有内容
        return open_excel
    def get_sheet_data(self,index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel().sheetnames#拿到sheetnames的所有内容
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data
    def get_cell_value(self,row,cols):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data().cell(row=row,column=cols)
        return data
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row
    def get_rows_value(self,row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

if __name__ == '__main__':
    handle = Excel()
    print(handle.get_rows_value(2))