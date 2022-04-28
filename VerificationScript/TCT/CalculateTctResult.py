from openpyxl import Workbook, load_workbook
import numpy as np
from scipy import interpolate


output_file_name = 'TCT result-2022-03-07 142657 - 副本.xlsx'
branch_index = [0, 1]
temperature_range_trx = range(-400, 1100, 100)
wb = load_workbook(output_file_name)
ws_tx = wb['Tx']
ws_Rx = wb['Rx']

ws_tx_calculation = wb.create_sheet("Tx_Calculation")
ws_rx_calculation = wb.create_sheet("Rx_Calculation")



for b in branch_index:
    cur_row=3
    temp = []
    tx_power = []

    while True:
        if ws_tx.cell(cur_row, b*4+1).value:
            temp.append(ws_tx.cell(cur_row, b*4+1).value)
            tx_power.append(ws_tx.cell(cur_row, b*4+2).value)
            cur_row = cur_row + 1
        else:
            cur_row = cur_row +3
            x_old = np.array(temp)
            x_old = x_old *10
            y_old = np.array(tx_power)
            x_new = np.array(temperature_range_trx)
            f = interpolate.interp1d(x_old, y_old, kind="slinear")
            print(f(x_new))

            cur_row = 3
            break
