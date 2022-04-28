import time
import numpy as np
from matplotlib import pyplot as plt
from Lib.Driver.Instruments import PowerMeter, SignalGenerator
from Lib.TestMethod import OBUE, PathLoss, TRDC
from Lib.Driver.Ericsson import RRU
from Lib.Basic import LogConfig
from Lib.Driver.Ericsson.RUMA import TCA
from openpyxl import Workbook

import math

###############setup pamameter###################
Com = "COM12"
output_file_name = 'TCT result-{}.xlsx'.format(time.strftime("%Y-%m-%d %H%M%S", time.localtime()))

branches = ['A', 'B']
rx_ipwr_FB = [0, 12]
tx_path_loss = -40
rx_path_loss = -40
tx_freq = 772e6
rx_freq = 717e6
mpa_index = [0, 1]
mpa_target = [350, 300]
mpa_start_offset = [1500, 1000]
dpa_index = [0, 1, 2, 3]
dpa_target = [75, 25, 25, 75]
dpa_start_offset = [3000, 3000, 3000, 3000]

temperature_range_trx = range(-400, 1100, 100)
temperature_norm_trx = 300
temperature_range_pa = range(-40, 110, 10)
temperature_norm_pa = 30
pa_prefix = '/pa/table'
trx_prefix = '/7xx/id_ruType_1.12.x/Band28F_68'
time_step = 1
time_total = 600
trdc_on_delay = 3
address_power_meter = 'GPIB0::9::INSTR'
address_signal_gen = 'TCPIP0::192.168.0.11::inst0::INSTR'
vdd_offset = ['rvc w FinalVdd:A 47950', 'rvc w FinalVdd:B 47950',
              'rvc w DriverVdd:0 47675']  # Need to set VDD to highest power class, we can read from ru but it is complicated, so just write it directly

############init test###################
pm = PowerMeter.PowerMeter(address_power_meter, timeout=10)
pm.reset()
pm.set_dbm()
'''
sg = SignalGenerator.SignalGenerator(address_signal_gen, timeout=10)
'''
# set sg later, currently load ARB waveform manually
#sg.reset()
'''
sg.set_freq(rx_freq)
'''
# sg.set_waveform('E_UTRA_UL_5')
branch_index = range(len(branches))
tct_size = len(temperature_range_trx)
LogConfig.init_logger()
testlog = LogConfig.get_logger()
tca = TCA()

ru = RRU.DutSerial(portx=Com, bps=115200, timex=5)
ru.openCom()
trdc_tx = TRDC.TRDC_DL(branches=[0], rru=ru, tca=tca, carriers_setup_string=['L50tm3p1_{}_P-3.0dB'.format(tx_freq)])

###set TCT to 0###
tx_zero_tct_xy = ''
rx_zero_TCT_x = ''
rx_zero_TCT_y = ''
for t in temperature_range_trx:
    tx_zero_tct_xy = tx_zero_tct_xy + '{} 0 '.format(t)
    rx_zero_TCT_x = rx_zero_TCT_x + '{} '.format(t)
    rx_zero_TCT_y = rx_zero_TCT_y + '0 '.format(t)

rx_temp = []
rx_highgain_noatt = []
rx_highgain_att = []
rx_bypass_noatt = []
rx_bypass_att = []
tor_temp = []
tor_outputpower = []
pa_final_temp = []
pa_driver_temp = []
mpa_offset = []
dpa_offset = []

# set all trx TCT to 0
for i in branch_index:

    ru.sendWait('db resize {}/tx{}/tor/tempCorTab {}'.format(trx_prefix, branches[i], tct_size * 2))
    ru.sendWait('db resize {}/rx1{}/highGain/noAtt/gainTempCompTable_x {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize {}/rx1{}/highGain/noAtt/gainTempCompTable_y {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx1{}/highGain/att/gainTempCompTable_x {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx1{}/highGain/att/gainTempCompTable_y {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx1{}/bypass/noAtt/gainTempCompTable_x {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx1{}/bypass/noAtt/gainTempCompTable_y {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx1{}/bypass/att/gainTempCompTable_x {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx1{}/bypass/att/gainTempCompTable_y {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx3{}/gainTempCompTable_x {}'.format(trx_prefix, branches[i], tct_size))
    ru.sendWait('db resize /stub{}/rx3{}/gainTempCompTable_y {}'.format(trx_prefix, branches[i], tct_size))

    ru.sendWait('db write {}/tx{}/tor/tempCorTab {}'.format(trx_prefix, branches[i], tx_zero_tct_xy))
    ru.sendWait('db write {}/rx1{}/highGain/noAtt/gainTempCompTable_x {}'.format(trx_prefix, branches[i], rx_zero_TCT_x))
    ru.sendWait('db write {}/rx1{}/highGain/noAtt/gainTempCompTable_y {}'.format(trx_prefix, branches[i], rx_zero_TCT_y))
    ru.sendWait('db write {}/rx1{}/highGain/att/gainTempCompTable_x {}'.format(trx_prefix, branches[i], rx_zero_TCT_x))
    ru.sendWait('db write {}/rx1{}/highGain/att/gainTempCompTable_y {}'.format(trx_prefix, branches[i], rx_zero_TCT_y))
    ru.sendWait('db write {}/rx1{}/bypass/noAtt/gainTempCompTable_x {}'.format(trx_prefix, branches[i], rx_zero_TCT_x))
    ru.sendWait('db write {}/rx1{}/bypass/noAtt/gainTempCompTable_y {}'.format(trx_prefix, branches[i], rx_zero_TCT_y))
    ru.sendWait('db write {}/rx1{}/bypass/att/gainTempCompTable_x {}'.format(trx_prefix, branches[i], rx_zero_TCT_x))
    ru.sendWait('db write {}/rx1{}/bypass/att/gainTempCompTable_y {}'.format(trx_prefix, branches[i], rx_zero_TCT_y))
    ru.sendWait('db write {}/rx3{}/gainTempCompTable_x {}'.format(trx_prefix, branches[i], rx_zero_TCT_x))
    ru.sendWait('db write {}/rx3{}/gainTempCompTable_y {}'.format(trx_prefix, branches[i], rx_zero_TCT_y))

    rx_temp.append([])
    rx_highgain_noatt.append([])
    rx_highgain_att.append([])
    rx_bypass_noatt.append([])
    rx_bypass_att.append([])
    tor_temp.append([])
    tor_outputpower.append([])
    pa_final_temp.append([])
    pa_driver_temp.append([])
    mpa_offset.append([])
    dpa_offset.append([])
    for j in range(len(mpa_index)):
        pa_final_temp[i].append([])
        mpa_offset[i].append([])

    for j in range(len(dpa_index)):
        pa_driver_temp[i].append([])
        dpa_offset[i].append([])

    # setup Rx carrier
    ru.sendWait(
        'trdc setup7 {} 1 0 0 {} 0 0 0 0 1 {} 0 0 1 0 11 1 {} {} 4 0 '.format(len(branches) + i + 1, int(rx_freq / 1e3),
                                                                              i, math.floor(i / 2), i % 2))
    ru.sendWait('trdc on {}'.format(len(branches) + i + 1))

start_time = time.time()
while time.time() - start_time < time_total:

    for i in branch_index:
        #####read rx temp and ipwr
        rx_temp[i].append(ru.get_single_data('ts r Rx:{}'.format(branches[i])))
        #sg.set_rf_state('ON')
        ru.set_rx_front_end(branch=branches[i], bypass='off', att='off')
        time.sleep(0.5)
        rx_highgain_noatt[i].append(ru.get_rxipwr_fb(fb = rx_ipwr_FB[i], repeat = 10))
        ru.set_rx_front_end(branch=branches[i], bypass='off', att='on')
        time.sleep(0.5)
        rx_highgain_att[i].append(ru.get_rxipwr_fb(fb = rx_ipwr_FB[i], repeat = 10))
        ru.set_rx_front_end(branch=branches[i], bypass='on', att='off')
        time.sleep(0.5)
        rx_bypass_noatt[i].append(ru.get_rxipwr_fb(fb = rx_ipwr_FB[i], repeat = 10))
        ru.set_rx_front_end(branch=branches[i], bypass='on', att='on')
        time.sleep(0.5)
        rx_bypass_att[i].append(ru.get_rxipwr_fb(fb = rx_ipwr_FB[i], repeat = 10))
        #sg.set_rf_state('OFF')

    ##########PA cal##########################
    for i in vdd_offset:
        ru.sendWait(i)   #set VDD to highest power class
    for b in branch_index:
        ru.sendWait('sw s PaOn:{} on'.format(branches[b]))
        ####final PA#####

        for i in range(len(mpa_index)):
            mpa_offset[b][i].append(ru.calibrate_pa(name='Mpa:{}.{}'.format(b, mpa_index[i]), target=mpa_target[i],start_offset=mpa_start_offset[i]))
            pa_final_temp[b][i].append(ru.get_single_data('ts r PaFinal:{}'.format(branches[b])))
        ####driver PA#####

        for i in range(len(dpa_index)):
            dpa_offset[b][i].append(ru.calibrate_pa(name='Dpa:{}.{}'.format(b, dpa_index[i]), target=dpa_target[i],start_offset=dpa_start_offset[i]))
            pa_driver_temp[b][i].append(ru.get_single_data('ts r PaDriver:{}'.format(branches[b])))

        ru.sendWait('sw s PaOn:{} off'.format(branches[b]))

    ####read Tx temp and power####

    for b in branch_index:
        trdc_tx.branches = [b]
        trdc_tx.del_all_carrier()
        trdc_tx.set_up_DL_carrier_init()
        time.sleep(5)
        tor_outputpower[b].append(pm.read_power())
        tor_temp[b].append(ru.get_single_data('ts r Tor:{}'.format(branches[b])))
        trdc_tx.trdc_release_all()

    time.sleep(time_step)


############save data#############
wb = Workbook()
wb.remove(wb.active)
##########write PA to excel###############
ws_pa = wb.create_sheet("PA")
dpa_col_offs = len(branches) * len(mpa_index) * 3
for b in branch_index:
    for i in range(len(mpa_index)):
        ws_pa.cell(row=1, column=b * len(mpa_index) * 3 + i * 3 + 1).value = 'MPA {}.{}'.format(b, i)
        ws_pa.cell(row=2, column=b*len(mpa_index) * 3 + i * 3 + 1).value = 'Temp'
        ws_pa.cell(row=2, column=b*len(mpa_index) * 3 + i * 3 + 2).value = 'Offset'
        for j in range(len(pa_final_temp[b][i])):
            ws_pa.cell(row=j + 3, column=b * len(mpa_index) * 3 + i * 3 + 1).value = pa_final_temp[b][i][j]
            ws_pa.cell(row=j + 3, column=b * len(mpa_index) * 3 + i * 3 + 2).value = mpa_offset[b][i][j]

    for i in range(len(dpa_index)):
        ws_pa.cell(row=1, column=dpa_col_offs + b*len(dpa_index) * 3 + i * 3 + 1).value = 'DPA {}.{}'.format(b, i)
        ws_pa.cell(row=2, column=dpa_col_offs + b*len(dpa_index) * 3 + i * 3 + 1).value = 'Temp'
        ws_pa.cell(row=2, column=dpa_col_offs + b*len(dpa_index) * 3 + i * 3 + 2).value = 'Offset'
        for j in range(len(pa_driver_temp[b][i])):
            ws_pa.cell(row=j + 3, column=dpa_col_offs + b * len(dpa_index) * 3 + i * 3 + 1).value = pa_driver_temp[b][i][j]
            ws_pa.cell(row=j + 3, column=dpa_col_offs + b * len(dpa_index) * 3 + i * 3 + 2).value = dpa_offset[b][i][j]


##########write TX to excel###############
ws_tx = wb.create_sheet("Tx")
for b in branch_index:
    ws_tx.cell(row=1, column=b * 3 + 1).value = 'Tx {}'.format(b)
    ws_tx.cell(row=2, column=b * 3 + 1).value = 'Temp'
    ws_tx.cell(row=2, column=b * 3 + 2).value = 'Power'
    for j in range(len(tor_temp[b])):
        ws_tx.cell(row=j + 3, column=b * 3 + 1).value = tor_temp[b][j]
        ws_tx.cell(row=j + 3, column=b * 3 + 2).value = tor_outputpower[b][j]

##########write RX to excel###############
ws_tx = wb.create_sheet("Rx")
for b in branch_index:
    ws_tx.cell(row=1, column=b * 6 + 1).value = 'Rx {}'.format(b)
    ws_tx.cell(row=2, column=b * 6 + 1).value = 'Temp'
    ws_tx.cell(row=2, column=b * 6 + 2).value = 'HighGainNoAtt'
    ws_tx.cell(row=2, column=b * 6 + 3).value = 'HighGainAtt'
    ws_tx.cell(row=2, column=b * 6 + 4).value = 'BypassNoAtt'
    ws_tx.cell(row=2, column=b * 6 + 5).value = 'BypassAtt'
    for j in range(len(rx_temp[b])):
        ws_tx.cell(row=j + 3, column=b * 6 + 1).value = rx_temp[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 2).value = rx_highgain_noatt[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 3).value = rx_highgain_att[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 4).value = rx_bypass_noatt[b][j]
        ws_tx.cell(row=j + 3, column=b * 6 + 5).value = rx_bypass_att[b][j]

wb.save(output_file_name)


for b in branch_index:
    ru.sendWait('trdc release {}'.format(len(branches) + b + 1))
ru.closeCom()
