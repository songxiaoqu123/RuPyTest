import time
from Lib.Basic import LogConfig
from Lib.Driver.Ericsson import RRU
from Lib.Driver import Instruments
from Lib.Driver.Instruments import SignalGenerator
from Lib.Driver.Instruments import VSA_Keysight as VectorSpectrumAnalyser
from openpyxl import Workbook

LogConfig.init_logger()
testlog = LogConfig.get_logger()
from Lib.Driver.SwitchBox import HRSM
from Lib.Driver.Ericsson.RUMA import TCA
from Lib.TestMethod import PathLoss, TRDC

########Tset Setup######################################################
dut_com = ['COM10', 'COM11']
cpriPorts = ['1A', '1B']
port = ['A', 'B', 'C', 'D']
output_file = 'HALT.xlsx'
total_time = 120

tx_norm = 47.8
tx_carriers_setup_string = ['NR200tm3p1a_2120e6_P-0dB']
tx_pathloss = [-37.0, -36.7, -37.1, -37, -37.0, -36.7, -37.1, -37]

rx_norm = -50
rx_gain_norm = 151.4

rx_carriers_setup_string = ['NR200tm3p1a_1930e6']
rx_pathloss = [-33.2, -33.0, -33.3, -33.1, -33.2, -33.0, -33.3, -33.1]
rx_pwr_index = [0, 6, 12, 18, 0, 6, 12, 18]

sg = SignalGenerator.SignalGenerator('TCPIP0::192.168.0.63::inst0::INSTR')
pm = None
vsa = VectorSpectrumAnalyser.VectorSpectrumAnalyzer('TCPIP0::192.168.0.22::hislip0::INSTR')

##########Init Test######################################################
dut_index = range(len(dut_com))
port_index = range(len(port))
switch_matrix = HRSM.HRSM()
wb = Workbook()
wb.remove(wb.active)
tca = TCA()

tx_power = 'NA'
tx_dev = 'NA'
rx_power = 'NA'
rx_dev = 'NA'


tx_freq = tx_carriers_setup_string[0].split('_')[1]
rx_freq = rx_carriers_setup_string[0].split('_')[1]

starttime = time.strftime("%Y%m%d%H%M%S", time.localtime())
output_file = starttime + output_file
current_sample = 0
outputfile = []
dut = []
ws = []
current_row = 1
current_col = 1
dut_key = ['time', 'Sample']
vsa_tap = ['Power', 'ACLR1_L', 'ACLR1_R', 'ACLR2_L', 'ACLR2_R', 'EVM_rms', 'Freq_err', 'IQ_offset']
port_key = []
port_result ={}
if pm:
    port_key.append('TxPower')
    port_key.append('TxDev')
if sg:
    port_key.append('RxPower')
    port_key.append('RxDev')
if vsa:
    for tap in vsa_tap:
        port_key.append(tap)
port_key.append('dpd_restart_c')
for key in port_key:
    port_result[key] = []

for i in dut_index:
    dut.append(RRU.DutSerial(portx=dut_com[i], bps=115200, timex=5))
    dut[i].openCom()
txpower_data = dut[0].get_trxpower()
for key in txpower_data[0:-2:2]:
    port_key.append(key)
    port_result[key] = []

for i in dut_index:
    sn = dut[i].par_get()['SYS_HW_SERIAL']
    ws.append(wb.create_sheet(sn))
    for d in dut_key:
        ws[i].cell(row=current_row, column=current_col).value = d
        current_col += 1

    for key in port_key:
        for j in port_index:
            ws[i].cell(row=current_row, column=current_col).value = '{}_{}'.format(key, port[j])
            current_col += 1

    ts_data = dut[i].get_ts()
    for d in ts_data[0::2]:
        ws[i].cell(row=current_row, column=current_col).value = 'ts' + d
        current_col += 1

    vs_data = dut[i].get_vs()
    for d in vs_data[0::2]:
        ws[i].cell(row=current_row, column=current_col).value = 'vs' + d
        current_col += 1

    cs_data = dut[i].get_cs()
    for d in cs_data[0::2]:
        ws[i].cell(row=current_row, column=current_col).value = 'cs' + d
        current_col += 1

    rvc_data = dut[i].get_rvc()
    for d in rvc_data[0::2]:
        ws[i].cell(row=current_row, column=current_col).value = 'rvc' + d
        current_col += 1

    ws[i].cell(row=current_row, column=current_col).value = 'fm_fault'
    current_col += 1

    current_col = 1
    wb.save(output_file)
current_row += 1

trdc_dl = []
trdc_ul = []
tca.IQFileClearAll()

for i in dut_index:
    trdc_dl.append(
        TRDC.TRDC_DL(branches=port_index, rru=dut[i], tca=tca, carriers_setup_string=tx_carriers_setup_string,
                     trdc_start_index=1, cpriPorts=cpriPorts[i]))
    trdc_ul.append(
        TRDC.TRDC_UL(branches=port_index, rru=dut[i], tca=tca, carriers_setup_string=rx_carriers_setup_string,
                     trdc_start_index=1 + len(port_index), cpriPorts=cpriPorts[i]))
    trdc_dl[i].del_all_carrier()
    trdc_dl[i].set_up_DL_carrier_init()
    trdc_dl[i].trdc_release_all()
    trdc_ul[i].del_all_carrier()
    trdc_ul[i].set_up_UL_carrier_init()
    for j in port_index:
        dut[i].sendWait(r'trdc setup7 {} 4 6 1 {} 0 0 -0 0 1 {} -0 0 1 0 1 1 0 0 4 0'.format(j+1, int(float(tx_freq)/1e3),j))
        dut[i].sendWait(r'trdc on {}'.format(j+1))

#########StartTest###################################################
starttime = time.time()
try:
    while time.time() - starttime < total_time:
        for i in dut_index:
            ws[i].cell(row=current_row, column=current_col).value = time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime())
            current_col += 1
            ws[i].cell(row=current_row, column=current_col).value = current_sample
            current_col += 1

            for j in port_index:
                test_index = i * len(port) + j
                switch_matrix.set_path(switch_matrix.path1[test_index])
                time.sleep(1)
                if pm:
                    tx_power = pm.read_power() - tx_pathloss[test_index]
                    port_result['TxPower'].append(tx_power)
                    port_result['TxDev'].append(tx_power - tx_norm)


                if vsa:
                    vsa.send('*RCL 1')
                    vsa.set_freq_center(tx_freq)
                    vsa.set_ext_gain_BTS(tx_pathloss[test_index])
                    acp_data = vsa.get_acp()
                    for key in acp_data.keys():
                        if key in port_result.keys():
                            port_result[key].append(acp_data[key])
                    vsa.send('*RCL 2')
                    vsa.set_freq_center(tx_freq)
                    vsa.set_ext_gain_BTS(tx_pathloss[test_index])
                    evm_data = vsa.get_evm()
                    for key in evm_data.keys():
                        if key in port_result.keys():
                            port_result[key].append(acp_data[key])

                switch_matrix.set_path(switch_matrix.path2[test_index])
                if sg:
                    sg.set_freq(rx_freq)
                    sg.set_amp(rx_norm - rx_pathloss[test_index])
                    sg.set_rf_state('ON')
                    time.sleep(0.5)
                    rx_power = dut[i].get_rxipwr()
                    sg.set_rf_state('OFF')
                    rx_power = rx_power[rx_pwr_index[test_index]]
                    rx_power = rx_power - rx_gain_norm
                    port_result['RxPower'].append(rx_power)
                    port_result['RxDev'].append(rx_power - rx_norm)


                txpower_data = dut[i].get_trxpower(port=j)
                for d in txpower_data[1:-1:2]:
                    ws[i].cell(row=current_row, column=current_col).value = d
                    current_col += 1

                restart_count = dut[i].get_dpdRestartCount(port=port_index[j])
                ws[i].cell(row=current_row, column=current_col).value = restart_count
                current_col += 1

            ts_data = dut[i].get_ts()
            for d in ts_data[1::2]:
                ws[i].cell(row=current_row, column=current_col).value = d
                current_col += 1

            vs_data = dut[i].get_vs()
            for d in vs_data[1::2]:
                ws[i].cell(row=current_row, column=current_col).value = d
                current_col += 1

            cs_data = dut[i].get_cs()
            for d in cs_data[1::2]:
                ws[i].cell(row=current_row, column=current_col).value = d
                current_col += 1

            rvc_data = dut[i].get_rvc()
            for d in rvc_data[1::2]:
                ws[i].cell(row=current_row, column=current_col).value = d
                current_col += 1

            fm_fault = dut[i].get_fmfault()
            if fm_fault == 'No raised fault was found':
                fm_fault = 0
            ws[i].cell(row=current_row, column=current_col).value = fm_fault
            current_col += 1

            current_col = 1
        current_row += 1
        current_sample += 1
        wb.save(output_file)
except:
    for i in dut_index:
        trdc_dl[i].trdc_release_all()
        trdc_ul[i].trdc_release_all()
        dut[i].closeCom()
for i in dut_index:
    trdc_dl[i].trdc_release_all()
    trdc_ul[i].trdc_release_all()
    dut[i].closeCom()

wb.save(output_file)
